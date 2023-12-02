from openpyxl import load_workbook
from flask import Flask, render_template, request,redirect, url_for, flash,session
from flask_session import Session
app = Flask(__name__)
@app.route('/')
def index():
    return render_template('log.html')  # 假设register.html文件与此Python文件在同一目录下
@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
       # 这里可以处理用户的注册信息，例如保存到数据库等
        username = request.form['username']
        password = request.form['password']
        workbook = load_workbook("zqw.xlsx")
        sheet3 = workbook["注册"]
        sheet1 = workbook["配置"]
        sheet2 = workbook["名字"]
        sheet = workbook["全盟考勤"]
   #     app.secret_key = 'supersecretkey'
    #    app.config['SECRET_KEY'] = 'fileststem'
    #    Session(app)
    # 创建一个set来存储所有用户名
        usernames = {row[0] for row in sheet3.iter_rows(min_row=2, values_only=True)}
        if username in usernames:
            return "此名字已被人注册使用"
    # 插入新的用户名和密码
        sheet3.insert_rows(2)
        sheet1.insert_rows(2)
        sheet2.insert_rows(2)
        sheet.insert_rows(2)
        sheet3.cell(row=2, column=1, value=username)
        sheet.cell(row=2, column=1, value=username)
        sheet1.cell(row=2, column=1, value=username)
        sheet2.cell(row=2, column=1, value=username)
        sheet1.cell(row=2, column=2, value="5")
        sheet3.cell(row=2, column=2, value=password)
    # 假设其他sheets的操作与上面相似，省略以简化示例
        workbook.save('zqw.xlsx')
        return redirect(url_for('index'))
    return render_template('reg.html')

if __name__ == '__main__':
    app.run(debug=True)