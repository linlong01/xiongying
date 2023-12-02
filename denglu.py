from flask import Flask, render_template, request, redirect, url_for, flash,session
from openpyxl import load_workbook
from flask_session import Session
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Replace this with a proper secret key
@app.route('/')
def index():
    return render_template('log.html')
@app.route('/validate', methods=['POST'])
def validate():
    workbook = load_workbook("zqw.xlsx")
    sheet = workbook["注册"]
    input_username = request.form['username']
    input_password = request.form['password']
    user_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[0]
        password = row[1]
        if input_username == username:
            user_found = True
            if input_password == password:
                session['username'] = input_username
                return redirect(url_for('table'))
            else:
                flash("密码错误")
                return redirect(url_for('index'))
    if not user_found:
        flash("用户名不存在")
        return redirect(url_for('index'))
    return redirect(url_for('index')) # Added this as a fallback 
@app.route('/table')
def table():
    username = session.get('username','Guest')
    return render_template('yunx.html',username = username)
    #return render_template('log.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
       # 这里可以处理用户的注册信息，例如保存到数据库等
        return redirect(url_for('index'))  # 或者可以重定向到其他页面，如登录页面
    return render_template('reg.html')
if __name__ == '__main__':
    app.run(debug=True)