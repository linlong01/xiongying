from openpyxl import load_workbook
from flask import Flask, render_template, request,redirect, url_for, flash,session
from flask_session import Session  
import pandas as pd
import xlrd
import datetime
app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
@app.route('/')
def index():
    return render_template('log.html')
@app.route('/mingzi')
def mingzi():
    username = session.get('username','Guest')
    return render_template('yunx.html',username = username)
  #  if 'username' in Session :
  #      name = Session['username']
  #      return f"当前用户：Session['username']"
@app.route('/qiandao', methods=['GET','POST'])
def qiandao():
    workface = xlrd.open_workbook("zqw.xlsx")
    file_path = 'zqw.xlsx'
    workbook = load_workbook("zqw.xlsx")
    data = pd.read_excel(file_path , header=0)
    sheet = workbook["全盟考勤"]
    sheet1 = workbook["配置"]
    sheet2 = workbook["名字"]
    workSheet = workface.sheet_by_name("全盟考勤")
    data_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = {'请假': row[1], '缺勤': row[2], '积分': row[3]}
        data_dict[key] = value
    data_row = session.get('username','Guest') #"雄鹰"#input("游戏名字:")  # 取该行数据
    data_col = '积分'   # 取该列数据
    rowIndex = workSheet.col_values(0).index(data_row)
#ph = sheet1.cell(1,1)
#ph.value = int(ph.value)
    while True :
   # a1 = ph.value
        d_time = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '12:0', '%Y-%m-%d%H:%M')
        d_time1 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '14:00', '%Y-%m-%d%H:%M')
        d_time3 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '21:00', '%Y-%m-%d%H:%M')
        d_time4 = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '23:00', '%Y-%m-%d%H:%M')
        n_time = datetime.datetime.now()
        if n_time > d_time and n_time < d_time1 or n_time > d_time3 and n_time < d_time4:
            b1 = rowIndex + 1
            cell1 = sheet2.cell(b1,2)
            if cell1.value == "1" :
                return"您在此时段已打过卡"
               # break
            else :
                ui = sheet1.cell(b1,2)
                ui.value = int(ui.value)
                cell = sheet.cell(b1,ui.value)
                if data_row in data_dict :
                    de = sheet.cell(b1,3)
                    hg = sheet.cell(b1,4)
                    de.value = int(de.value)
                    hg.value = int(hg.value)
                    pi = request.form['zhu']#input("主力数：")
                    pt = request.form['chai']#input("拆迁数：")
                    if pi == "0" and pt == "0":
                        cell.value = "false"
                        cell1.value = "1"
                        de.value = de.value + 1
                        ui.value = ui.value + 1
                        workbook.save("zqw.xlsx")
                    elif pi == "0" or pt == "0":
                        cell.value = "考勤成功"
                        cell1.value = "1"
                        hg.value = hg.value+1
                        ui.value = ui.value + 1
                        workbook.save("zqw.xlsx")
                    else :
                        cell.value = "考勤成功"
                        cell1.value = "1"
                        hg.value = hg.value+2
                        ui.value = ui.value + 1
                        workbook.save("zqw.xlsx")
                   # break
                    return"完成"
          #  break
            return"报错"

        else :
            print("当前不是打卡时间")
            sheet2.delete_cols(2)
            workbook.save("zqw.xlsx")
            return"当前不是打卡时间"
          #  break
if __name__ == '__main__':
    app.run(debug=True)